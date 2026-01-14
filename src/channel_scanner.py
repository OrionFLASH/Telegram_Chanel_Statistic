"""
Модуль для сканирования каналов Telegram.

Содержит класс ChannelScanner для получения информации о всех каналах,
на которые подписан пользователь или в которых он является участником.
"""

import asyncio
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from telethon import TelegramClient
from telethon.tl import functions
from telethon.errors import ChatAdminRequiredError, FloodWaitError
from telethon.tl.types import Channel, Chat, User

from logger_config import get_logger


class ChannelScanner:
    """
    Класс для сканирования и сбора информации о каналах Telegram.
    
    Позволяет получить полный список всех каналов, групп и супергрупп,
    в которых пользователь является участником, с детальной информацией.
    """
    
    def __init__(
        self,
        client: TelegramClient,
        concurrency: int = 16,
        request_delay: float = 0.2,
        unsubscribe_ids: Optional[Set[int]] = None,
    ) -> None:
        """
        Инициализация сканера каналов.
        
        Args:
            client: Экземпляр TelegramClient для работы с API
            concurrency: Максимальное количество одновременных запросов
            request_delay: Небольшая задержка между запросами для снижения нагрузки
            unsubscribe_ids: Набор ID каналов/групп для авто-отписки
        """
        self.client = client
        self.logger = get_logger("channel_scanner")
        self.channels_data: List[Dict[str, Any]] = []
        self.private_chats_data: List[Dict[str, Any]] = []
        self.concurrency = max(1, concurrency)
        self.request_delay = max(0.0, request_delay)
        self.output_dir = Path(__file__).parent.parent / "OUT"
        self.output_dir.mkdir(exist_ok=True)
        self.unsubscribe_ids = unsubscribe_ids or set()

    def _format_participants_count(self, participants_count: Any) -> str:
        """
        Приводит количество участников к строке для логов и вывода.
        
        Args:
            participants_count: Значение количества участников
        
        Returns:
            Строковое представление количества участников
        """
        if participants_count is None:
            return "Неизвестно"
        return str(participants_count)
    
    async def _fetch_participants_count(
        self,
        entity: Channel,
        full_channel_info: Optional[Any] = None,
        full_chat_info: Optional[Any] = None,
    ) -> Optional[Any]:
        """
        Пытается получить количество участников разными способами.
        
        Args:
            entity: Объект канала/группы из Telegram API
            full_channel_info: Полная информация о канале (если уже получена)
            full_chat_info: Полная информация о чате (если уже получена)
        
        Returns:
            Количество участников или None
        """
        if hasattr(entity, "participants_count") and entity.participants_count:
            return entity.participants_count
        try:
            if full_channel_info and hasattr(full_channel_info, "full_chat"):
                if hasattr(full_channel_info.full_chat, "participants_count"):
                    return full_channel_info.full_chat.participants_count
            if full_chat_info and hasattr(full_chat_info, "full_chat"):
                if hasattr(full_chat_info.full_chat, "participants_count"):
                    return full_chat_info.full_chat.participants_count
            if isinstance(entity, Channel):
                full_info = await self.client(functions.channels.GetFullChannelRequest(channel=entity))
                if hasattr(full_info, "full_chat") and hasattr(full_info.full_chat, "participants_count"):
                    return full_info.full_chat.participants_count
            if isinstance(entity, Chat):
                full_info = await self.client(functions.messages.GetFullChatRequest(chat_id=entity.id))
                if hasattr(full_info, "full_chat") and hasattr(full_info.full_chat, "participants_count"):
                    return full_info.full_chat.participants_count
        except ChatAdminRequiredError:
            return "Требуются права администратора"
        except Exception as e:
            self.logger.debug(f"Ошибка при получении количества участников через Full* методы: {e}")
        return None

    async def _fetch_linked_channel_info(self, linked_chat_id: int) -> Dict[str, Optional[str]]:
        """
        Получает информацию о связанном канале, если он доступен.
        
        Args:
            linked_chat_id: ID связанного канала
        
        Returns:
            Словарь с данными связанного канала
        """
        linked_data: Dict[str, Optional[str]] = {
            "linked_chat_id": str(linked_chat_id),
            "linked_chat_title": None,
            "linked_chat_username": None,
            "linked_chat_link": None,
        }
        try:
            linked_entity = await self.client.get_entity(linked_chat_id)
            linked_data["linked_chat_title"] = getattr(linked_entity, "title", None)
            linked_data["linked_chat_username"] = getattr(linked_entity, "username", None)
            if linked_data["linked_chat_username"]:
                linked_data["linked_chat_link"] = f"https://t.me/{linked_data['linked_chat_username']}"
        except Exception as e:
            self.logger.debug(f"Не удалось получить данные связанного канала {linked_chat_id}: {e}")
        return linked_data

    async def _fetch_forum_topics(self, entity: Channel, limit: int = 100) -> List[str]:
        """
        Получает список тем форума для супергруппы с включенными темами.
        
        Args:
            entity: Сущность супергруппы
            limit: Максимальное количество тем для выборки
        
        Returns:
            Список названий тем форума
        """
        topics: List[str] = []
        seen_titles: Set[str] = set()
        try:
            offset_id = 0
            offset_topic = 0
            remaining = limit
            while remaining > 0:
                result = await self.client(
                    functions.channels.GetForumTopicsRequest(
                        channel=entity,
                        q="",
                        offset_date=None,
                        offset_id=offset_id,
                        offset_topic=offset_topic,
                        limit=min(remaining, 100),
                    )
                )
                result_topics = getattr(result, "topics", [])
                if not result_topics:
                    break
                for topic in result_topics:
                    title = getattr(topic, "title", None)
                    if title and title not in seen_titles:
                        seen_titles.add(title)
                        topics.append(title)
                last_topic = result_topics[-1]
                offset_topic = getattr(last_topic, "id", 0) or 0
                offset_id = getattr(last_topic, "top_message", 0) or 0
                remaining = limit - len(topics)
        except Exception as e:
            self.logger.debug(f"Не удалось получить темы форума для {entity.id}: {e}")
        return topics

    async def _fetch_last_message_date(self, entity: Channel) -> Optional[str]:
        """
        Получает дату последнего сообщения в канале/чате/группе.
        
        Args:
            entity: Сущность канала/группы
        
        Returns:
            Дата последнего сообщения в ISO формате или None
        """
        try:
            messages = await self.client.get_messages(entity, limit=1)
            if messages and messages[0] and messages[0].date:
                return messages[0].date.isoformat()
        except Exception as e:
            self.logger.debug(f"Не удалось получить дату последнего сообщения для {entity.id}: {e}")
        return None

    async def _leave_channel_or_chat(self, entity: Channel) -> bool:
        """
        Выполняет отписку от канала или выход из группы.
        
        Args:
            entity: Сущность канала/группы
        
        Returns:
            True, если отписка выполнена успешно
        """
        try:
            if isinstance(entity, Channel):
                await self.client(functions.channels.LeaveChannelRequest(channel=entity))
            elif isinstance(entity, Chat):
                await self.client(functions.messages.DeleteChatUser(chat_id=entity.id, user_id="me"))
            return True
        except Exception as e:
            self.logger.error(f"Ошибка при попытке отписки от {entity.id}: {e}")
            return False

    async def get_channel_info(
        self,
        entity: Channel,
        last_message_date: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        Получает детальную информацию о канале.
        
        Args:
            entity: Объект канала из Telegram API
            last_message_date: Дата последнего сообщения (если уже получена)
        
        Returns:
            Словарь с информацией о канале или None в случае ошибки
        """
        try:
            self.logger.debug(f"Получение информации о канале: {entity.title}")
            
            # Получаем базовую информацию о канале
            await self.client.get_entity(entity)
            
            # Базовые данные канала
            channel_data: Dict[str, Any] = {
                "id": entity.id,
                "title": entity.title or "Без названия",
                "username": entity.username or "Нет username",
                "is_broadcast": entity.broadcast,  # True для каналов, False для групп
                "is_megagroup": entity.megagroup,  # True для супергрупп
                "is_gigagroup": getattr(entity, 'gigagroup', False),
                "access_hash": str(entity.access_hash) if hasattr(entity, 'access_hash') else None,
                "scanned_at": datetime.now().isoformat()
            }
            
            # Пытаемся получить полную информацию для расширенных данных
            full_channel_info = None
            full_chat_info = None
            try:
                if isinstance(entity, Channel):
                    full_channel_info = await self.client(
                        functions.channels.GetFullChannelRequest(channel=entity)
                    )
                elif isinstance(entity, Chat):
                    full_chat_info = await self.client(
                        functions.messages.GetFullChatRequest(chat_id=entity.id)
                    )
            except ChatAdminRequiredError:
                self.logger.debug("Требуются права администратора для получения полной информации")
            except Exception as e:
                self.logger.debug(f"Не удалось получить полную информацию: {e}")
            
            # Пытаемся получить количество участников
            participants_count = await self._fetch_participants_count(
                entity,
                full_channel_info=full_channel_info,
                full_chat_info=full_chat_info,
            )
            if participants_count is None:
                # Дополнительная попытка через подсчет участников (только для групп)
                if not entity.broadcast:
                    try:
                            count = 0
                            async for _ in self.client.iter_participants(entity):
                                count += 1
                                if count >= 10000:  # Ограничение для производительности
                                    break
                        participants_count = count if count < 10000 else ">10000"
                        except Exception as e:
                        self.logger.debug(f"Не удалось получить количество участников через iter_participants: {e}")
            channel_data["participants_count"] = participants_count

            # Связанный канал (если настроен)
            linked_chat_id = None
            if full_channel_info and hasattr(full_channel_info, "full_chat"):
                linked_chat_id = getattr(full_channel_info.full_chat, "linked_chat_id", None)
            if linked_chat_id:
                linked_data = await self._fetch_linked_channel_info(linked_chat_id)
                channel_data.update(linked_data)
            else:
                channel_data.update(
                    {
                        "linked_chat_id": None,
                        "linked_chat_title": None,
                        "linked_chat_username": None,
                        "linked_chat_link": None,
                    }
                )

            # Темы форума (если супергруппа с включенными темами)
            forum_topics: List[str] = []
            is_forum = bool(getattr(entity, "forum", False))
            if full_channel_info and hasattr(full_channel_info, "full_chat"):
                is_forum = is_forum or bool(getattr(full_channel_info.full_chat, "forum", False))
            if is_forum and isinstance(entity, Channel):
                forum_topics = await self._fetch_forum_topics(entity)
            channel_data["forum_topics"] = forum_topics
            channel_data["forum_topics_count"] = len(forum_topics)

            # Дата последнего сообщения
            if last_message_date:
                channel_data["last_message_date"] = last_message_date
            else:
                channel_data["last_message_date"] = await self._fetch_last_message_date(entity)
            
            # Дополнительная информация
            if hasattr(entity, 'about'):
                channel_data["about"] = entity.about or "Нет описания"
            else:
                channel_data["about"] = "Нет описания"
            
            # Дата создания (если доступна)
            if hasattr(entity, 'date'):
                channel_data["created_date"] = entity.date.isoformat() if entity.date else None
            else:
                channel_data["created_date"] = None
            
            # Проверяем, является ли канал публичным
            channel_data["is_public"] = entity.username is not None
            
            # Ссылка на канал
            if entity.username:
                channel_data["link"] = f"https://t.me/{entity.username}"
            else:
                channel_data["link"] = f"tg://resolve?domain={entity.id}"
            
            participants_text = self._format_participants_count(channel_data.get("participants_count"))
            self.logger.info(
                f"Успешно получена информация о канале: {channel_data['title']} "
                f"(подписчиков: {participants_text})"
            )
            return channel_data
            
        except FloodWaitError as e:
            self.logger.warning(f"Превышен лимит запросов. Ожидание {e.seconds} секунд")
            await asyncio.sleep(e.seconds)
            return await self.get_channel_info(entity)
        except Exception as e:
            self.logger.error(f"Ошибка при получении информации о канале {entity.title}: {e}")
            return None
    
    async def scan_all_channels(self) -> List[Dict[str, Any]]:
        """
        Сканирует все каналы, группы и супергруппы пользователя.
        
        Returns:
            Список словарей с информацией о каждом канале
        """
        self.logger.info("Начало сканирования каналов")
        self.channels_data = []
        
        try:
            # Получаем все диалоги (чаты, каналы, группы)
            self.logger.debug("Получение списка всех диалогов")
            dialogs = await self.client.get_dialogs()
            
            self.logger.info(f"Найдено диалогов: {len(dialogs)}")
            
            # Фильтруем только каналы и группы
            channels_and_groups = []
            last_message_map: Dict[int, Optional[str]] = {}
            for dialog in dialogs:
                entity = dialog.entity
                # Проверяем, является ли это каналом или группой
                if isinstance(entity, Channel):
                    channels_and_groups.append(entity)
                    message_date = None
                    if getattr(dialog, "message", None) and getattr(dialog.message, "date", None):
                        message_date = dialog.message.date.isoformat()
                    last_message_map[entity.id] = message_date
                    self.logger.debug(f"Найден канал/группа: {entity.title}")
            
            self.logger.info(f"Найдено каналов и групп: {len(channels_and_groups)}")
            
            # Получаем информацию о каждом канале с ограничением параллелизма
            semaphore = asyncio.Semaphore(self.concurrency)

            async def process_channel(index: int, entity: Channel) -> Optional[Dict[str, Any]]:
                """
                Обрабатывает один канал с ограничением параллелизма.
                
                Args:
                    index: Порядковый номер канала
                    entity: Сущность канала/группы
                
                Returns:
                    Словарь с информацией о канале или None
                """
                async with semaphore:
                    self.logger.info(
                        f"Обработка канала {index}/{len(channels_and_groups)}: {entity.title}"
                    )
                    channel_info = await self.get_channel_info(
                        entity,
                        last_message_date=last_message_map.get(entity.id),
                    )
                if channel_info:
                        participants_text = self._format_participants_count(
                            channel_info.get("participants_count")
                        )
                        self.logger.info(
                            f"Канал обработан: {channel_info.get('title', '')} "
                            f"(подписчиков: {participants_text})"
                        )
                        if channel_info.get("id") in self.unsubscribe_ids:
                            self.logger.info(
                                f"Отписка по списку: {channel_info.get('title', '')} "
                                f"(id: {channel_info.get('id')})"
                            )
                            is_unsubscribed = await self._leave_channel_or_chat(entity)
                            if is_unsubscribed:
                                channel_info["unsubscribed_status"] = "Да"
                            else:
                                channel_info["unsubscribed_status"] = "Ошибка отписки"
                        else:
                            channel_info["unsubscribed_status"] = "Нет"
                    # Небольшая задержка между запросами
                    if self.request_delay:
                        await asyncio.sleep(self.request_delay)
                    return channel_info

            tasks = [
                process_channel(index, channel)
                for index, channel in enumerate(channels_and_groups, 1)
            ]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for result in results:
                if isinstance(result, dict):
                    self.channels_data.append(result)
                elif isinstance(result, Exception):
                    self.logger.error(f"Ошибка при обработке канала: {result}")
            
            self.logger.info(f"Сканирование завершено. Обработано каналов: {len(self.channels_data)}")
            return self.channels_data
            
        except Exception as e:
            self.logger.error(f"Критическая ошибка при сканировании каналов: {e}")
            raise
    
    def save_to_json(self, filename: str = "channels_data.json") -> None:
        """
        Сохраняет данные о каналах в JSON файл.
        
        Args:
            filename: Имя файла для сохранения
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_path = self.output_dir / self._append_timestamp(filename, timestamp)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.channels_data, f, ensure_ascii=False, indent=2)
            self.logger.info(f"Данные сохранены в файл: {output_path}")
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении данных: {e}")
            raise
    
    def save_to_text(self, filename: str = "channels_list.txt") -> None:
        """
        Сохраняет данные о каналах в текстовый файл для удобного чтения.
        
        Args:
            filename: Имя файла для сохранения
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_path = self.output_dir / self._append_timestamp(filename, timestamp)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("СПИСОК КАНАЛОВ И ГРУПП TELEGRAM\n")
                f.write(f"Дата сканирования: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Всего найдено: {len(self.channels_data)}\n")
                f.write("=" * 80 + "\n\n")
                
                for i, channel in enumerate(self.channels_data, 1):
                    f.write(f"\n{'=' * 80}\n")
                    f.write(f"Канал #{i}\n")
                    f.write(f"{'=' * 80}\n")
                    f.write(f"Название: {channel['title']}\n")
                    f.write(f"ID: {channel['id']}\n")
                    f.write(f"Username: {channel['username']}\n")
                    f.write(f"Тип: ")
                    if channel['is_broadcast']:
                        f.write("Канал (Broadcast)\n")
                    elif channel['is_megagroup']:
                        f.write("Супергруппа (Megagroup)\n")
                    elif channel['is_gigagroup']:
                        f.write("Гигагруппа (Gigagroup)\n")
                    else:
                        f.write("Группа\n")
                    f.write(f"Публичный: {'Да' if channel['is_public'] else 'Нет'}\n")
                    f.write(f"Количество участников: {channel.get('participants_count', 'Неизвестно')}\n")
                    f.write(f"Описание: {channel.get('about', 'Нет описания')}\n")
                    f.write(f"Ссылка: {channel['link']}\n")
                    if channel.get('created_date'):
                        f.write(f"Дата создания: {channel['created_date']}\n")
                    f.write(f"Дата сканирования: {channel['scanned_at']}\n")
                
            self.logger.info(f"Текстовый отчет сохранен в файл: {output_path}")
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении текстового файла: {e}")
            raise

    def _build_xlsx_rows(self) -> Tuple[List[str], List[List[str]]]:
        """
        Подготавливает заголовки и строки для выгрузки в XLSX.
        
        Returns:
            Заголовки и строки в виде списков строк
        """
        headers = [
            "ID",
            "Название",
            "Username",
            "Тип",
            "Публичный",
            "Участников",
            "Описание",
            "Ссылка",
            "Связанный канал ID",
            "Связанный канал",
            "Связанный канал ссылка",
            "Удален по списку",
            "Темы форума (кол-во)",
            "Темы форума",
            "Дата создания",
            "Дата последнего сообщения",
            "Дата сканирования",
        ]
        rows: List[List[Any]] = []
        sorted_channels = sorted(
            self.channels_data,
            key=self._participants_sort_key,
            reverse=True,
        )
        for channel in sorted_channels:
            if channel.get("is_broadcast"):
                channel_type = "Канал"
            elif channel.get("is_megagroup"):
                channel_type = "Супергруппа"
            elif channel.get("is_gigagroup"):
                channel_type = "Гигагруппа"
            else:
                channel_type = "Группа"
            participants_value = channel.get("participants_count")
            if isinstance(participants_value, int):
                participants_cell: Any = participants_value
            elif isinstance(participants_value, str) and participants_value.isdigit():
                participants_cell = int(participants_value)
            elif participants_value is None:
                participants_cell = "Неизвестно"
            else:
                participants_cell = str(participants_value)
            forum_topics = channel.get("forum_topics") or []
            forum_topics_text = "; ".join(str(item) for item in forum_topics) if forum_topics else ""
            rows.append(
                [
                    str(channel.get("id", "")),
                    str(channel.get("title", "")),
                    str(channel.get("username", "")),
                    channel_type,
                    "Да" if channel.get("is_public") else "Нет",
                    participants_cell,
                    str(channel.get("about", "")),
                    str(channel.get("link", "")),
                    str(channel.get("linked_chat_id", "")) if channel.get("linked_chat_id") else "",
                    str(channel.get("linked_chat_title", "")) if channel.get("linked_chat_title") else "",
                    str(channel.get("linked_chat_link", "")) if channel.get("linked_chat_link") else "",
                    str(channel.get("unsubscribed_status", "")),
                    int(channel.get("forum_topics_count", 0) or 0),
                    forum_topics_text,
                    str(channel.get("created_date", "")) if channel.get("created_date") else "",
                    str(channel.get("last_message_date", "")) if channel.get("last_message_date") else "",
                    str(channel.get("scanned_at", "")),
                ]
            )
        return headers, rows

    def _participants_sort_key(self, channel: Dict[str, Any]) -> int:
        """
        Возвращает числовой ключ для сортировки по количеству участников.
        
        Args:
            channel: Данные канала
        
        Returns:
            Число участников, если доступно, иначе 0
        """
        participants_value = channel.get("participants_count")
        if isinstance(participants_value, int):
            return participants_value
        if isinstance(participants_value, str) and participants_value.isdigit():
            return int(participants_value)
        return 0

    def _apply_xlsx_styles(
        self,
        worksheet,
        headers: List[str],
        rows: List[List[Any]],
        numeric_formats: Optional[Dict[str, str]] = None,
    ) -> None:
        """
        Применяет форматирование к листу XLSX для удобного просмотра.
        
        Args:
            worksheet: Лист Excel
            headers: Список заголовков
            rows: Строки данных
            numeric_formats: Форматы для числовых колонок по названию
        """
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        data_alignment = Alignment(vertical="top", wrap_text=True)
        numeric_alignment = Alignment(horizontal="right", vertical="top")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        zebra_fill = PatternFill(start_color="F3F6FA", end_color="F3F6FA", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        numeric_formats = numeric_formats or {}
        numeric_cols = {
            headers.index(name) + 1: fmt
            for name, fmt in numeric_formats.items()
            if name in headers
        }
        for row_idx, row in enumerate(rows, start=2):
            is_zebra = row_idx % 2 == 0
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in numeric_cols:
                    cell.alignment = numeric_alignment
                    if isinstance(value, (int, float)):
                        cell.number_format = numeric_cols[col_idx]
                else:
                    cell.alignment = data_alignment
                cell.border = thin_border
                if is_zebra:
                    cell.fill = zebra_fill

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

        for col_idx, header in enumerate(headers, 1):
            if header == "Темы форума":
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 80
                continue
            max_len = len(header)
            for row in rows:
                value = row[col_idx - 1]
                if value:
                    max_len = max(max_len, len(str(value)))
            adjusted = min(max(max_len + 2, 12), 60)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted

    def save_to_xlsx(self, filename: str = "channels_data.xlsx") -> None:
        """
        Сохраняет данные о каналах в XLSX файл с удобным форматированием.
        
        Args:
            filename: Имя файла для сохранения
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_path = self.output_dir / self._append_timestamp(filename, timestamp)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Каналы"

            headers, rows = self._build_xlsx_rows()
            worksheet.append(headers)
            for row in rows:
                worksheet.append(row)

            self._apply_xlsx_styles(
                worksheet,
                headers,
                rows,
                numeric_formats={
                    "Участников": "#,##0",
                    "Темы форума (кол-во)": "#,##0",
                },
            )

            private_headers, private_rows = self._build_private_xlsx_rows()
            private_sheet = workbook.create_sheet(title="Личные чаты")
            private_sheet.append(private_headers)
            for row in private_rows:
                private_sheet.append(row)
            self._apply_xlsx_styles(
                private_sheet,
                private_headers,
                private_rows,
                numeric_formats={
                    "Сообщений за 90 дней": "#,##0",
                    "Среднее в день": "#,##0.00",
                    "Среднее в неделю": "#,##0.00",
                    "Среднее в месяц": "#,##0.00",
                    "Сообщений за год": "#,##0",
                    "Сообщений всего": "#,##0",
                    "Сообщений от вас": "#,##0",
                    "Сообщений от собеседника": "#,##0",
                },
            )
            workbook.save(output_path)
            self.logger.info(f"XLSX отчет сохранен в файл: {output_path}")
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении XLSX файла: {e}")
            raise

    def _append_timestamp(self, filename: str, timestamp: str) -> str:
        """
        Добавляет таймштамп к имени файла перед расширением.
        
        Args:
            filename: Имя файла
            timestamp: Таймштамп в формате YYYYMMDD_HHMM
        
        Returns:
            Имя файла с добавленным таймштампом
        """
        if "." in filename:
            name, ext = filename.rsplit(".", 1)
            return f"{name}_{timestamp}.{ext}"
        return f"{filename}_{timestamp}"

    async def scan_private_chats(self) -> List[Dict[str, Any]]:
        """
        Сканирует личные чаты и собирает статистику сообщений.
        
        Returns:
            Список словарей с информацией о личных чатах
        """
        self.logger.info("Начало сканирования личных чатов")
        self.private_chats_data = []
        try:
            dialogs = await self.client.get_dialogs()
            private_dialogs = []
            last_message_map: Dict[int, Optional[str]] = {}
            for dialog in dialogs:
                entity = dialog.entity
                if isinstance(entity, User) and not getattr(entity, "bot", False):
                    if getattr(entity, "is_self", False):
                        continue
                    private_dialogs.append(entity)
                    message_date = None
                    if getattr(dialog, "message", None) and getattr(dialog.message, "date", None):
                        message_date = dialog.message.date.isoformat()
                    last_message_map[entity.id] = message_date

            self.logger.info(f"Найдено личных чатов: {len(private_dialogs)}")
            semaphore = asyncio.Semaphore(self.concurrency)

            async def process_private_chat(index: int, entity: User) -> Optional[Dict[str, Any]]:
                """
                Обрабатывает личный чат с ограничением параллелизма.
                
                Args:
                    index: Порядковый номер чата
                    entity: Пользователь чата
                
                Returns:
                    Словарь с данными личного чата или None
                """
                async with semaphore:
                    self.logger.info(
                        f"Обработка личного чата {index}/{len(private_dialogs)}: "
                        f"{getattr(entity, 'first_name', '')} {getattr(entity, 'last_name', '')}".strip()
                    )
                    chat_info = await self._collect_private_chat_info(
                        entity,
                        last_message_map.get(entity.id),
                    )
                    if self.request_delay:
                        await asyncio.sleep(self.request_delay)
                    return chat_info

            tasks = [
                process_private_chat(index, entity)
                for index, entity in enumerate(private_dialogs, 1)
            ]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for result in results:
                if isinstance(result, dict):
                    self.private_chats_data.append(result)
                elif isinstance(result, Exception):
                    self.logger.error(f"Ошибка при обработке личного чата: {result}")
            return self.private_chats_data
        except Exception as e:
            self.logger.error(f"Критическая ошибка при сканировании личных чатов: {e}")
            raise

    async def _collect_private_chat_info(
        self,
        entity: User,
        last_message_date: Optional[str],
    ) -> Optional[Dict[str, Any]]:
        """
        Собирает информацию и статистику по личному чату.
        
        Args:
            entity: Пользователь
            last_message_date: Дата последнего сообщения (если уже известна)
        
        Returns:
            Словарь с данными личного чата
        """
        try:
            display_name = " ".join(
                part for part in [getattr(entity, "first_name", ""), getattr(entity, "last_name", "")] if part
            ).strip() or "Без имени"
            now = datetime.now(timezone.utc)
            threshold_90 = now - timedelta(days=90)
            threshold_year = now - timedelta(days=365)
            messages_90 = 0
            messages_year = 0
            messages_total = 0
            messages_from_me = 0
            messages_from_other = 0

            async for message in self.client.iter_messages(entity):
                if not getattr(message, "date", None):
                    continue
                message_date = message.date
                messages_total += 1
                if message.out:
                    messages_from_me += 1
                else:
                    messages_from_other += 1
                if message_date >= threshold_year:
                    messages_year += 1
                if message_date >= threshold_90:
                    messages_90 += 1

            average_per_day = round(messages_90 / 90, 2)
            average_per_week = round(messages_90 / (90 / 7), 2)
            average_per_month = round(messages_90 / 3, 2)

            return {
                "id": entity.id,
                "name": display_name,
                "username": getattr(entity, "username", None),
                "phone": getattr(entity, "phone", None),
                "participants": f"Вы; {display_name}",
                "last_message_date": last_message_date,
                "messages_90": messages_90,
                "avg_day": average_per_day,
                "avg_week": average_per_week,
                "avg_month": average_per_month,
                "messages_year": messages_year,
                "messages_total": messages_total,
                "messages_from_me": messages_from_me,
                "messages_from_other": messages_from_other,
            }
        except Exception as e:
            self.logger.error(f"Ошибка при сборе данных личного чата {entity.id}: {e}")
            return None

    def _build_private_xlsx_rows(self) -> Tuple[List[str], List[List[Any]]]:
        """
        Формирует данные для листа с личными чатами.
        
        Returns:
            Заголовки и строки для XLSX
        """
        headers = [
            "ID",
            "Имя",
            "Username",
            "Телефон",
            "Участники",
            "Дата последнего сообщения",
            "Сообщений за 90 дней",
            "Среднее в день",
            "Среднее в неделю",
            "Среднее в месяц",
            "Сообщений за год",
            "Сообщений всего",
            "Сообщений от вас",
            "Сообщений от собеседника",
        ]
        rows: List[List[Any]] = []
        for chat in self.private_chats_data:
            rows.append(
                [
                    str(chat.get("id", "")),
                    str(chat.get("name", "")),
                    str(chat.get("username", "")) if chat.get("username") else "",
                    str(chat.get("phone", "")) if chat.get("phone") else "",
                    str(chat.get("participants", "")),
                    str(chat.get("last_message_date", "")) if chat.get("last_message_date") else "",
                    int(chat.get("messages_90", 0)),
                    float(chat.get("avg_day", 0.0)),
                    float(chat.get("avg_week", 0.0)),
                    float(chat.get("avg_month", 0.0)),
                    int(chat.get("messages_year", 0)),
                    int(chat.get("messages_total", 0)),
                    int(chat.get("messages_from_me", 0)),
                    int(chat.get("messages_from_other", 0)),
                ]
            )
        return headers, rows
